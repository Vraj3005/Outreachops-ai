import csv
import hashlib
import io
import json
import logging
import os
from typing import Any

from openpyxl import load_workbook

from app.database import supabase

logger = logging.getLogger("outreachops.services.import")

# Standard target fields
STANDARD_FIELDS = [
    "first_name",
    "last_name",
    "full_name",
    "company_name",
    "job_title",
    "contact_email",
    "phone",
    "website",
    "industry",
    "country",
    "city",
    "tags",
]


class ImportService:
    def __init__(self):
        self.scratch_dir = os.path.join(
            os.path.dirname(os.path.dirname(__file__)), "scratch"
        )
        os.makedirs(self.scratch_dir, exist_ok=True)

    def calculate_fingerprint(self, contents: bytes) -> str:
        """Generates MD5 hash for idempotency and caching."""
        return hashlib.md5(contents).hexdigest()

    def get_cache_path(self, fingerprint: str) -> str:
        return os.path.join(self.scratch_dir, f"temp_import_{fingerprint}.json")

    def cleanup_expired_cache(self):
        """Clean up temp JSON files in scratch directory older than 1 hour."""
        import time

        now = time.time()
        one_hour_sec = 3600
        try:
            if os.path.exists(self.scratch_dir):
                for filename in os.listdir(self.scratch_dir):
                    if filename.startswith("temp_import_") and filename.endswith(
                        ".json"
                    ):
                        filepath = os.path.join(self.scratch_dir, filename)
                        file_mod_time = os.path.getmtime(filepath)
                        if (now - file_mod_time) > one_hour_sec:
                            try:
                                os.remove(filepath)
                                logger.info(
                                    f"Cleaned up expired import cache file: {filename}"
                                )
                            except Exception:
                                pass
        except Exception as e:
            logger.warning(f"Error cleaning up import cache files: {e}")

    def parse_file(self, contents: bytes, file_name: str) -> dict[str, Any]:
        """
        Parses CSV or XLSX bytes. Limits size, columns, rows, and cell length.
        Saves parsed raw data into scratch JSON file named by MD5 fingerprint.
        """
        self.cleanup_expired_cache()

        # Limit file size < 10MB
        if len(contents) > 10 * 1024 * 1024:
            raise ValueError("File size exceeds maximum limit of 10MB.")

        fingerprint = self.calculate_fingerprint(contents)
        cache_path = self.get_cache_path(fingerprint)

        # If already cached, return metadata immediately
        if os.path.exists(cache_path):
            with open(cache_path, encoding="utf-8") as f:
                cached = json.load(f)
            return {
                "fingerprint": fingerprint,
                "headers": cached["headers"],
                "sample_rows": cached["rows"][:5],
                "total_rows": len(cached["rows"]),
            }

        headers: list[str] = []
        rows: list[list[str]] = []

        ext = os.path.splitext(file_name.lower())[1]

        if ext in [".xlsx", ".xls"]:
            # XLSX Parsing using read-only to prevent zip bombs
            try:
                # data_only=True ignores raw formulas and returns pre-calculated values
                wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
                sheet = wb.active
                if not sheet:
                    raise ValueError("Workbook has no active sheets.")

                # Read rows streaming
                row_idx = 0
                for row_cells in sheet.iter_rows(values_only=True):
                    row_idx += 1
                    # Limit rows < 5000
                    if row_idx > 5000:
                        break

                    # Convert to string list, checking cell lengths
                    cell_vals: list[str] = []
                    for cell in row_cells:
                        val = "" if cell is None else str(cell)
                        if len(val) > 5000:
                            raise ValueError(
                                f"Row {row_idx}: Cell length exceeds limit of 5000 characters."
                            )
                        cell_vals.append(val.strip())

                    # Stop if entirely empty row
                    if not any(cell_vals):
                        continue

                    if row_idx == 1:
                        headers = cell_vals
                        # Limit columns < 100
                        if len(headers) > 100:
                            raise ValueError(
                                "File columns exceed maximum limit of 100."
                            )
                    else:
                        rows.append(cell_vals)
                wb.close()
            except Exception as e:
                logger.error(f"XLSX Parsing failure: {e}")
                if "limit" in str(e).lower() or "exceeds" in str(e).lower():
                    raise e
                raise ValueError("XLSX parsing failed. Verify file validity.")
        else:
            # CSV Parsing with encoding fallbacks
            decoded = ""
            for encoding in ["utf-8-sig", "latin-1", "utf-8"]:
                try:
                    decoded = contents.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue

            if not decoded:
                raise ValueError(
                    "Could not decode file content with supported encodings."
                )

            csv_file = io.StringIO(decoded)
            reader = csv.reader(csv_file)

            try:
                row_idx = 0
                for r in reader:
                    row_idx += 1
                    if row_idx > 5000:
                        break

                    # Cell length limit checks
                    cell_vals = []
                    for c in r:
                        if len(c) > 5000:
                            raise ValueError(
                                f"Row {row_idx}: Cell length exceeds limit of 5000 characters."
                            )
                        cell_vals.append(c.strip())

                    if not any(cell_vals):
                        continue

                    if row_idx == 1:
                        headers = cell_vals
                        if len(headers) > 100:
                            raise ValueError(
                                "File columns exceed maximum limit of 100."
                            )
                    else:
                        rows.append(cell_vals)
            except Exception as e:
                logger.error(f"CSV Parsing failure: {e}")
                if "limit" in str(e).lower() or "exceeds" in str(e).lower():
                    raise e
                raise ValueError("CSV parsing failed. Verify file format.")

        # Ensure we have data
        if not headers:
            raise ValueError("Parsed file headers are empty.")

        # Handle duplicate headers by appending indices
        seen_headers = {}
        unique_headers = []
        for h in headers:
            h_clean = h.strip() or "column"
            if h_clean in seen_headers:
                seen_headers[h_clean] += 1
                unique_headers.append(f"{h_clean}_{seen_headers[h_clean]}")
            else:
                seen_headers[h_clean] = 1
                unique_headers.append(h_clean)

        # Cache data in temporary scratch file
        cache_data = {"headers": unique_headers, "rows": rows}
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache_data, f, ensure_ascii=False, indent=2)

        return {
            "fingerprint": fingerprint,
            "headers": unique_headers,
            "sample_rows": rows[:5],
            "total_rows": len(rows),
        }

    def parse_google_sheet_rows(self, rows_data: list[list[Any]]) -> dict[str, Any]:
        """Formats spreadsheet rows identically to parsed files."""
        if not rows_data:
            raise ValueError("Spreadsheet data is empty.")

        headers: list[str] = [str(h).strip() for h in rows_data[0]]
        if len(headers) > 100:
            raise ValueError("Spreadsheet columns exceed maximum limit of 100.")

        rows: list[list[str]] = []
        for idx, r in enumerate(rows_data[1:], start=2):
            if idx > 5000:
                break
            cell_vals = []
            for cell in r:
                val = "" if cell is None else str(cell)
                if len(val) > 5000:
                    raise ValueError(
                        f"Row {idx}: Cell length exceeds limit of 5000 characters."
                    )
                cell_vals.append(val.strip())

            if not any(cell_vals):
                continue
            rows.append(cell_vals)

        # Seen headers deduplication
        seen_headers = {}
        unique_headers = []
        for h in headers:
            h_clean = h or "column"
            if h_clean in seen_headers:
                seen_headers[h_clean] += 1
                unique_headers.append(f"{h_clean}_{seen_headers[h_clean]}")
            else:
                seen_headers[h_clean] = 1
                unique_headers.append(h_clean)

        # Generate a fingerprint hash for the sheet content
        data_serialized = json.dumps({"headers": unique_headers, "rows": rows})
        fingerprint = hashlib.md5(data_serialized.encode("utf-8")).hexdigest()

        # Save cache
        cache_path = self.get_cache_path(fingerprint)
        with open(cache_path, "w", encoding="utf-8") as f:
            f.write(data_serialized)

        return {
            "fingerprint": fingerprint,
            "headers": unique_headers,
            "sample_rows": rows[:5],
            "total_rows": len(rows),
        }

    def auto_suggest_mappings(
        self, headers: list[str], mapping_preset: dict[str, str] = None
    ) -> dict[str, str]:
        """
        Maps parsed headers to target columns.
        Uses explicit mapping preset configuration or falls back to fuzzy match.
        """
        suggestions: dict[str, str] = {}
        preset = mapping_preset or {}

        for h in headers:
            h_lower = h.lower().replace("_", "").replace(" ", "").replace("-", "")

            # Check presets first
            if h in preset:
                suggestions[h] = preset[h]
                continue
            elif h_lower in preset:
                suggestions[h] = preset[h_lower]
                continue

            # Case-insensitive checks
            matched = False
            for tf in STANDARD_FIELDS:
                tf_clean = tf.replace("_", "")
                if h_lower == tf_clean:
                    suggestions[h] = tf
                    matched = True
                    break

            if matched:
                continue

            # Common heuristic aliases
            if h_lower in ["first", "fname"]:
                suggestions[h] = "first_name"
            elif h_lower in ["last", "lname"]:
                suggestions[h] = "last_name"
            elif h_lower in ["name", "fullname", "contactname"]:
                suggestions[h] = "full_name"
            elif h_lower in ["company", "companyname", "org", "organization"]:
                suggestions[h] = "company_name"
            elif h_lower in ["title", "jobtitle", "role"]:
                suggestions[h] = "job_title"
            elif h_lower in ["email", "contactemail", "emailaddress", "mail"]:
                suggestions[h] = "contact_email"
            elif h_lower in ["phone", "phonenumber", "tel", "cell"]:
                suggestions[h] = "phone"
            elif h_lower in ["website", "domain", "url", "site"]:
                suggestions[h] = "website"
            elif h_lower in ["industry", "sector"]:
                suggestions[h] = "industry"
            elif h_lower in ["country", "nation"]:
                suggestions[h] = "country"
            elif h_lower in ["city", "town", "locality"]:
                suggestions[h] = "city"
            elif h_lower in ["tags", "tag", "labels"]:
                suggestions[h] = "tags"
            else:
                # Normalize unmatched headers as custom fields
                normal_name = h.lower().strip().replace(" ", "_").replace("-", "_")
                normal_name = "".join(c for c in normal_name if c.isalnum() or c == "_")
                suggestions[h] = f"custom_fields.{normal_name}"

        return suggestions

    def validate_records(
        self,
        fingerprint: str,
        field_mapping: dict[str, str],
        user_id: str,
        duplicate_keys: list[str] = None,
    ) -> dict[str, Any]:
        """
        Loads the cached import file. Mapped target values are validated.
        Checks for row duplicates and existing DB contacts.
        """
        cache_path = self.get_cache_path(fingerprint)
        if not os.path.exists(cache_path):
            raise ValueError("Import payload session has expired. Please re-upload.")

        with open(cache_path, encoding="utf-8") as f:
            cached = json.load(f)

        headers = cached["headers"]
        rows = cached["rows"]

        valid_records: list[dict[str, Any]] = []
        row_logs: list[dict[str, Any]] = []

        seen_payload_duplicates = set()

        for idx, row in enumerate(rows, start=2):
            errors = []
            mapped = {}
            customs = {}

            # Map raw cell values based on configuration
            for col_idx, cell in enumerate(row):
                if col_idx >= len(headers):
                    continue
                header = headers[col_idx]
                target = field_mapping.get(header)
                if not target:
                    continue

                if target.startswith("custom_fields."):
                    field_key = target.split(".", 1)[1]
                    customs[field_key] = cell
                else:
                    mapped[target] = cell

            # Fill missing target keys with empty strings
            website = mapped.get("website", "").strip()
            email = mapped.get("contact_email", "").strip()

            # 1. Format Validations
            if not website:
                errors.append("Website domain is a required field.")
            elif "." not in website or len(website) < 3:
                errors.append("Invalid website domain format.")

            if not email:
                errors.append("Contact email is a required field.")
            elif "@" not in email or "." not in email:
                errors.append("Invalid contact email format.")

            # 2. Duplicate Check inside current import payload
            dup_check_key = f"{website.lower()}|{email.lower()}"
            if dup_check_key in seen_payload_duplicates:
                errors.append("Duplicate lead row inside upload file.")
            else:
                if website and email:
                    seen_payload_duplicates.add(dup_check_key)

            # 3. Duplicate checks in the database (Only if not already faulted)
            if not errors and supabase:
                try:
                    db_check = (
                        supabase.table("leads").select("id").eq("user_id", user_id)
                    )

                    # Deduplicate using configurable keys (default: website + email)
                    db_check = db_check.eq("website", website).eq(
                        "contact_email", email
                    )

                    res = db_check.execute()
                    if res.data:
                        errors.append("Lead website & email already exist in database.")
                except Exception as e:
                    logger.error(f"Failed database duplicate validation check: {e}")

            is_valid = len(errors) == 0

            # Package mapped lead row properties
            full_record = {
                "first_name": mapped.get("first_name"),
                "last_name": mapped.get("last_name"),
                "full_name": mapped.get("full_name")
                or mapped.get("company_name", website.split(".")[0].capitalize()),
                "company_name": mapped.get("company_name")
                or website.split(".")[0].capitalize(),
                "job_title": mapped.get("job_title"),
                "contact_email": email,
                "phone": mapped.get("phone"),
                "website": website,
                "industry": mapped.get("industry"),
                "country": mapped.get("country"),
                "city": mapped.get("city"),
                "tags": (
                    [t.strip() for t in mapped.get("tags", "").split(",") if t.strip()]
                    if mapped.get("tags")
                    else []
                ),
                "custom_fields": customs,
                "source_row_number": str(idx),
            }

            if is_valid:
                valid_records.append(full_record)

            row_logs.append(
                {
                    "row_number": idx,
                    "is_valid": is_valid,
                    "errors": errors,
                    "preview_data": {
                        "website": website,
                        "email": email,
                        "company_name": full_record["company_name"],
                    },
                    "original_cells": row,
                }
            )

        # Save the validation row logs in cache for downloading error CSV files later
        log_cache_path = cache_path.replace(".json", "_validation_logs.json")
        with open(log_cache_path, "w", encoding="utf-8") as f:
            json.dump(row_logs, f, ensure_ascii=False, indent=2)

        return {
            "fingerprint": fingerprint,
            "total_rows": len(rows),
            "valid_count": len(valid_records),
            "error_count": len(rows) - len(valid_records),
            "preview_items": row_logs[:10],
            "errors_list": [log for log in row_logs if not log["is_valid"]],
        }

    def generate_error_csv(self, fingerprint: str) -> str:
        """Loads cached validation logs and returns CSV containing all error details."""
        cache_path = self.get_cache_path(fingerprint)
        log_cache_path = cache_path.replace(".json", "_validation_logs.json")

        if not os.path.exists(cache_path) or not os.path.exists(log_cache_path):
            raise ValueError("Error log session has expired. Re-run validation.")

        with open(cache_path, encoding="utf-8") as f:
            cached_file = json.load(f)
        headers = cached_file["headers"]

        with open(log_cache_path, encoding="utf-8") as f:
            row_logs = json.load(f)

        def escape_csv_val(val):
            if val is None:
                return ""
            val_str = str(val)
            if val_str and val_str[0] in ["=", "+", "-", "@"]:
                return "'" + val_str
            return val

        output = io.StringIO()
        writer = csv.writer(output)

        # Header + Error Details Column
        writer.writerow(["Row Number"] + headers + ["Validation Error Details"])

        for log in row_logs:
            if not log["is_valid"]:
                row_num = log["row_number"]
                err_text = "; ".join(log["errors"])
                cells = log["original_cells"]
                # Align values
                cells = list(cells) + [""] * max(0, len(headers) - len(cells))
                escaped_cells = [escape_csv_val(c) for c in cells]
                writer.writerow([row_num] + escaped_cells + [escape_csv_val(err_text)])

        return output.getvalue()

    def commit_import_records(
        self,
        fingerprint: str,
        field_mapping: dict[str, str],
        user_id: str,
        source_name: str,
        source_type: str,
        preset_name: str = None,
    ) -> dict[str, Any]:
        """
        Commits all validated valid records to Supabase under source configurations.
        Saves mapping presets if preset_name is supplied.
        """
        # 1. Validate and fetch valid records
        validation = self.validate_records(fingerprint, field_mapping, user_id)
        cache_path = self.get_cache_path(fingerprint)

        with open(cache_path, encoding="utf-8") as f:
            cached = json.load(f)

        # Resolve validation details
        log_cache_path = cache_path.replace(".json", "_validation_logs.json")
        with open(log_cache_path, encoding="utf-8") as f:
            row_logs = json.load(f)

        valid_records = []
        for log in row_logs:
            if log["is_valid"]:
                # Rebuild valid item
                mapped = {}
                customs = {}
                row = log["original_cells"]

                for col_idx, cell in enumerate(row):
                    if col_idx >= len(cached["headers"]):
                        continue
                    header = cached["headers"][col_idx]
                    target = field_mapping.get(header)
                    if not target:
                        continue

                    if target.startswith("custom_fields."):
                        field_key = target.split(".", 1)[1]
                        customs[field_key] = cell
                    else:
                        mapped[target] = cell

                website = mapped.get("website", "").strip()
                email = mapped.get("contact_email", "").strip()

                valid_records.append(
                    {
                        "user_id": user_id,
                        "first_name": mapped.get("first_name"),
                        "last_name": mapped.get("last_name"),
                        "full_name": mapped.get("full_name")
                        or mapped.get(
                            "company_name", website.split(".")[0].capitalize()
                        ),
                        "company_name": mapped.get("company_name")
                        or website.split(".")[0].capitalize(),
                        "job_title": mapped.get("job_title"),
                        "contact_email": email,
                        "phone": mapped.get("phone"),
                        "website": website,
                        "industry": mapped.get("industry"),
                        "country": mapped.get("country"),
                        "city": mapped.get("city"),
                        "tags": (
                            [
                                t.strip()
                                for t in mapped.get("tags", "").split(",")
                                if t.strip()
                            ]
                            if mapped.get("tags")
                            else []
                        ),
                        "custom_fields": customs,
                        "lead_status": "Pending",
                        "source_sheet_name": source_name,
                        "source_row_number": str(log["row_number"]),
                    }
                )

        if not valid_records:
            return {
                "imported": 0,
                "skipped_duplicates": validation["error_count"],
                "total": validation["total_rows"],
            }

        # 2. Save mapping preset if requested
        mapping_id = None
        if preset_name and supabase:
            try:
                preset_payload = {
                    "user_id": user_id,
                    "name": preset_name,
                    "source_headers": cached["headers"],
                    "field_mapping": field_mapping,
                }
                # Check if exists
                existing = (
                    supabase.table("import_mappings")
                    .select("id")
                    .eq("user_id", user_id)
                    .eq("name", preset_name)
                    .execute()
                )
                if existing.data:
                    mapping_id = existing.data[0]["id"]
                    supabase.table("import_mappings").update(preset_payload).eq(
                        "id", mapping_id
                    ).execute()
                else:
                    insert_res = (
                        supabase.table("import_mappings")
                        .insert(preset_payload)
                        .execute()
                    )
                    if insert_res.data:
                        mapping_id = insert_res.data[0]["id"]
            except Exception as e:
                logger.error(f"Failed to save mapping preset details: {e}")

        # 3. Create Import Source record
        source_id = None
        if supabase:
            try:
                source_payload = {
                    "user_id": user_id,
                    "source_type": source_type,
                    "name": source_name,
                    "total_rows": validation["total_rows"],
                    "successful_rows": len(valid_records),
                    "failed_rows": validation["error_count"],
                    "mapping_id": mapping_id,
                }
                src_res = (
                    supabase.table("import_sources").insert(source_payload).execute()
                )
                if src_res.data:
                    source_id = src_res.data[0]["id"]
            except Exception as e:
                logger.error(f"Failed to record import source metadata: {e}")

        # 4. Ingest Leads
        imported_count = 0
        if supabase:
            # Batch inserts in chunks of 100
            chunk_size = 100
            for i in range(0, len(valid_records), chunk_size):
                chunk = valid_records[i : i + chunk_size]
                # Attach source_id
                if source_id:
                    for item in chunk:
                        item["source_id"] = source_id
                try:
                    supabase.table("leads").insert(chunk).execute()
                    imported_count += len(chunk)
                except Exception as e:
                    logger.error(f"Batch lead insert failed on chunk index {i}: {e}")

        # Clean cache files
        try:
            if os.path.exists(cache_path):
                os.remove(cache_path)
            if os.path.exists(log_cache_path):
                os.remove(log_cache_path)
        except Exception:
            pass

        return {
            "imported": imported_count,
            "failed_rows": validation["error_count"],
            "total_processed": validation["total_rows"],
        }
